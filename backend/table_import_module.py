"""
GastroCore Table Import & Seed Module
=====================================
Persistente Tisch-Stammdaten + Kombinationen

ENDPOINTS:
- GET  /api/data-status              - System-Status mit Counts
- POST /api/admin/import/tables      - Excel Upload für Tische
- POST /api/admin/import/table-combinations - Excel Upload für Kombinationen
- POST /api/admin/seed/from-repo     - Seed aus /seed/ Ordner

ADDITIV - Keine Breaking Changes
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
from pathlib import Path
import uuid
import os
import io
import logging

# Core imports
from core.database import db
from core.auth import require_admin
from core.audit import create_audit_log

logger = logging.getLogger(__name__)

# Router
import_router = APIRouter(tags=["Import & Seed"])

# Paths
SEED_FOLDER = Path("/app/seed")
SEED_FOLDER.mkdir(parents=True, exist_ok=True)


# ============== MODELS ==============
class DataStatusResponse(BaseModel):
    build_id: str
    version: str
    database_type: str  # "external" or "local"
    database_warning: Optional[str]
    counts: Dict[str, int]
    last_import: Optional[Dict[str, Any]]


class ImportResult(BaseModel):
    success: bool
    created: int
    updated: int
    errors: int
    message: str


# ============== HELPERS ==============
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def get_mongo_type() -> tuple:
    """Check if using external or local MongoDB"""
    mongo_url = os.environ.get('MONGO_URL', '')
    if 'localhost' in mongo_url or '127.0.0.1' in mongo_url or not mongo_url:
        return "local", "⚠️ Lokale MongoDB - Daten gehen bei Container-Neustart verloren!"
    return "external", None


async def log_import(user_email: str, filename: str, collection: str, created: int, updated: int, errors: int):
    """Log import operation"""
    await db.import_logs.insert_one({
        "id": str(uuid.uuid4()),
        "timestamp": now_iso(),
        "user": user_email,
        "filename": filename,
        "collection": collection,
        "created": created,
        "updated": updated,
        "errors": errors
    })


# ============== ENDPOINTS ==============

@import_router.get("/api/data-status", response_model=DataStatusResponse)
async def get_data_status():
    """
    GET /api/data-status
    System health with collection counts
    """
    db_type, warning = get_mongo_type()
    
    # Get counts
    counts = {
        "tables": await db.tables.count_documents({"archived": {"$ne": True}}),
        "table_combinations": await db.table_combinations.count_documents({"archived": {"$ne": True}}),
        "reservations": await db.reservations.count_documents({"archived": {"$ne": True}}),
        "staff_members": await db.staff_members.count_documents({"archived": {"$ne": True}}),
        "events": await db.events.count_documents({"archived": {"$ne": True}}),
        "actions": await db.actions.count_documents({"archived": {"$ne": True}}),
        "users": await db.users.count_documents({"archived": {"$ne": True}}),
    }
    
    # Get last import
    last_import = await db.import_logs.find_one(sort=[("timestamp", -1)])
    last_import_data = None
    if last_import:
        last_import_data = {
            "timestamp": last_import.get("timestamp"),
            "collection": last_import.get("collection"),
            "user": last_import.get("user"),
            "created": last_import.get("created"),
            "updated": last_import.get("updated")
        }
    
    # Get build info
    import subprocess
    try:
        commit = subprocess.check_output(['git', 'rev-parse', '--short', 'HEAD'], cwd='/app').decode().strip()
    except:
        commit = "unknown"
    
    return {
        "build_id": f"{commit}-{datetime.now().strftime('%Y%m%d')}",
        "version": "3.0.0",
        "database_type": db_type,
        "database_warning": warning,
        "counts": counts,
        "last_import": last_import_data
    }


@import_router.post("/api/admin/import/tables", response_model=ImportResult)
async def import_tables(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    """
    POST /api/admin/import/tables
    Import tables from Excel file (Sheet: tables)
    Upsert by (table_number, area, subarea)
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Nur Excel-Dateien (.xlsx) erlaubt")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    
    # Find tables sheet
    sheet_name = 'tables' if 'tables' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    
    # Get headers
    headers = [str(c.value).lower() if c.value else f"col{i}" for i, c in enumerate(ws[1])]
    
    created = 0
    updated = 0
    errors = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        try:
            # Map columns
            data = dict(zip(headers, row))
            
            table_number = str(data.get('table_number', '')).strip()
            area = str(data.get('area', 'restaurant')).strip()
            subarea = data.get('subarea') or data.get('sub_area')
            if subarea:
                subarea = str(subarea).strip()
                if subarea == 'terrasse':
                    subarea = None
            
            seats = int(data.get('seats', 4))
            max_seats = int(data.get('max_seats', seats))
            combinable = str(data.get('combinable', 'true')).lower() == 'true'
            notes = data.get('notes')
            
            # Find existing
            existing = await db.tables.find_one({
                "table_number": table_number,
                "area": area
            })
            
            now = now_iso()
            
            if existing:
                await db.tables.update_one(
                    {"id": existing['id']},
                    {"$set": {
                        "sub_area": subarea,
                        "seats_default": seats,
                        "seats_max": max_seats,
                        "combinable": combinable,
                        "notes": notes,
                        "updated_at": now
                    }}
                )
                updated += 1
            else:
                await db.tables.insert_one({
                    "id": str(uuid.uuid4()),
                    "table_number": table_number,
                    "area": area,
                    "sub_area": subarea,
                    "seats_default": seats,
                    "seats_max": max_seats,
                    "combinable": combinable,
                    "combinable_with": [],
                    "notes": notes,
                    "active": True,
                    "fixed": False,
                    "position_x": None,
                    "position_y": None,
                    "created_at": now,
                    "updated_at": now,
                    "archived": False
                })
                created += 1
        except Exception as e:
            logger.error(f"Error importing table row: {e}")
            errors += 1
    
    # Log import
    await log_import(user.get('email', 'unknown'), file.filename, 'tables', created, updated, errors)
    
    # Audit
    await create_audit_log(
        actor={"id": user['id'], "email": user['email']},
        entity="tables",
        entity_id="bulk_import",
        action="import",
        after={"filename": file.filename, "created": created, "updated": updated}
    )
    
    return {
        "success": errors == 0,
        "created": created,
        "updated": updated,
        "errors": errors,
        "message": f"Import abgeschlossen: {created} neu, {updated} aktualisiert, {errors} Fehler"
    }


@import_router.post("/api/admin/import/table-combinations", response_model=ImportResult)
async def import_table_combinations(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    """
    POST /api/admin/import/table-combinations
    Import table combinations from Excel file
    Upsert by combo_id
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Nur Excel-Dateien (.xlsx) erlaubt")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    
    # Find combinations sheet
    sheet_name = 'combinations' if 'combinations' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    
    # Get headers
    headers = [str(c.value).lower() if c.value else f"col{i}" for i, c in enumerate(ws[1])]
    
    created = 0
    updated = 0
    errors = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        try:
            data = dict(zip(headers, row))
            
            combo_id = str(data.get('combo_id', '')).strip()
            subarea = str(data.get('subarea', '')).strip()
            tables_str = str(data.get('tables', '')).strip()
            target_capacity = int(data.get('target_capacity', 0)) if data.get('target_capacity') else 0
            notes = data.get('notes')
            
            # Parse tables list
            tables_list = [t.strip() for t in tables_str.split('+') if t.strip()]
            
            # Determine area from subarea
            if subarea in ['saal', 'wintergarten']:
                area = 'restaurant'
            else:
                area = subarea  # terrasse
            
            # Find existing
            existing = await db.table_combinations.find_one({"combo_id": combo_id})
            
            now = now_iso()
            
            combo_data = {
                "combo_id": combo_id,
                "area": area,
                "subarea": subarea,
                "tables": tables_list,
                "target_capacity": target_capacity,
                "notes": notes,
                "updated_at": now
            }
            
            if existing:
                await db.table_combinations.update_one(
                    {"id": existing['id']},
                    {"$set": combo_data}
                )
                updated += 1
            else:
                combo_data["id"] = str(uuid.uuid4())
                combo_data["created_at"] = now
                combo_data["archived"] = False
                combo_data["active"] = True
                await db.table_combinations.insert_one(combo_data)
                created += 1
        except Exception as e:
            logger.error(f"Error importing combination row: {e}")
            errors += 1
    
    # Log import
    await log_import(user.get('email', 'unknown'), file.filename, 'table_combinations', created, updated, errors)
    
    # Audit
    await create_audit_log(
        actor={"id": user['id'], "email": user['email']},
        entity="table_combinations",
        entity_id="bulk_import",
        action="import",
        after={"filename": file.filename, "created": created, "updated": updated}
    )
    
    return {
        "success": errors == 0,
        "created": created,
        "updated": updated,
        "errors": errors,
        "message": f"Import abgeschlossen: {created} neu, {updated} aktualisiert, {errors} Fehler"
    }


@import_router.post("/api/admin/seed/from-repo", response_model=ImportResult)
async def seed_from_repo(user: dict = Depends(require_admin)):
    """
    POST /api/admin/seed/from-repo
    Load tables and combinations from /seed/ folder
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    tables_file = SEED_FOLDER / "tables.xlsx"
    combos_file = SEED_FOLDER / "table_combinations.xlsx"
    
    if not tables_file.exists() and not combos_file.exists():
        raise HTTPException(404, "Keine Seed-Dateien gefunden in /seed/")
    
    total_created = 0
    total_updated = 0
    total_errors = 0
    
    # Import tables
    if tables_file.exists():


# ============== STAFF IMPORT ==============

import hashlib
import re

def normalize_phone(phone: str) -> str:
    """Normalize phone number to +49 format"""
    if not phone:
        return None
    phone = str(phone).strip()
    phone = re.sub(r'[^\d+]', '', phone)
    if phone.startswith('0') and not phone.startswith('00'):
        phone = '+49' + phone[1:]
    elif phone.startswith('49') and not phone.startswith('+'):
        phone = '+' + phone
    elif not phone.startswith('+'):
        phone = '+49' + phone
    return phone if len(phone) > 5 else None


def parse_date(date_val) -> str:
    """Parse date from various formats to ISO"""
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m-%d")
    try:
        # Try DD.MM.YYYY
        return datetime.strptime(str(date_val), "%d.%m.%Y").strftime("%Y-%m-%d")
    except:
        try:
            # Try YYYY-MM-DD
            return datetime.strptime(str(date_val), "%Y-%m-%d").strftime("%Y-%m-%d")
        except:
            return None


def create_row_hash(data: dict) -> str:
    """Create hash for duplicate detection"""
    key = f"{data.get('last_name','')}{data.get('first_name','')}{data.get('email','')}{data.get('phone','')}"
    return hashlib.md5(key.lower().encode()).hexdigest()[:16]


class StaffImportResult(BaseModel):
    success: bool
    imported: int
    updated: int
    skipped: int
    errors: int
    error_details: List[Dict[str, Any]]
    message: str


@import_router.post("/api/admin/import/staff", response_model=StaffImportResult)
async def import_staff(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    """
    POST /api/admin/import/staff
    Import staff members from Excel file
    Upsert by email or (first_name + last_name + personnel_number)
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Nur Excel-Dateien (.xlsx) erlaubt")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Get headers (normalize to lowercase)
    raw_headers = [str(c.value).lower().strip() if c.value else f"col{i}" for i, c in enumerate(ws[1])]
    
    # Column mapping
    col_map = {
        'nachname': 'last_name',
        'vorname': 'first_name',
        'rufname': 'display_name',
        'e-mail': 'email',
        'email': 'email',
        'telefon': 'phone',
        'telefon 2': 'phone_secondary',
        'adresse': 'street',
        'stadt': 'city',
        'plz': 'zip',
        'geburtstag': 'birthday',
        'pers-nr.': 'personnel_number',
        'pers-nr': 'personnel_number',
        'zeit-pin': 'time_pin',
    }
    
    imported = 0
    updated = 0
    skipped = 0
    errors = 0
    error_details = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        try:
            # Map row to dict
            raw_data = dict(zip(raw_headers, row))
            data = {}
            for excel_col, db_field in col_map.items():
                if excel_col in raw_data:
                    data[db_field] = raw_data[excel_col]
            
            # Validation
            first_name = str(data.get('first_name', '')).strip()
            last_name = str(data.get('last_name', '')).strip()
            
            if not first_name or not last_name:
                error_details.append({"row": row_idx, "reason": "Vorname oder Nachname fehlt"})
                errors += 1
                continue
            
            # Normalize fields
            email = str(data.get('email', '')).strip().lower() if data.get('email') else None
            phone = normalize_phone(data.get('phone'))
            phone_secondary = normalize_phone(data.get('phone_secondary'))
            birthday = parse_date(data.get('birthday'))
            personnel_number = str(data.get('personnel_number', '')).strip() if data.get('personnel_number') else None
            time_pin = str(data.get('time_pin', '')).strip() if data.get('time_pin') else None
            display_name = str(data.get('display_name', '')).strip() if data.get('display_name') else first_name
            
            # Address
            address = None
            if data.get('street') or data.get('city') or data.get('zip'):
                address = {
                    "street": str(data.get('street', '')).strip() if data.get('street') else None,
                    "city": str(data.get('city', '')).strip() if data.get('city') else None,
                    "zip": str(data.get('zip', '')).strip() if data.get('zip') else None,
                }
            
            # Create row hash for duplicate detection
            row_hash = create_row_hash({
                'last_name': last_name,
                'first_name': first_name,
                'email': email,
                'phone': phone
            })
            
            # Find existing - try multiple strategies
            existing = None
            
            # Strategy 1: by email
            if email:
                existing = await db.staff_members.find_one({"email": email, "archived": {"$ne": True}})
            
            # Strategy 2: by phone
            if not existing and phone:
                existing = await db.staff_members.find_one({"phone": phone, "archived": {"$ne": True}})
            
            # Strategy 3: by name + personnel_number
            if not existing and personnel_number:
                existing = await db.staff_members.find_one({
                    "first_name": first_name,
                    "last_name": last_name,
                    "personnel_number": personnel_number,
                    "archived": {"$ne": True}
                })
            
            # Strategy 4: by external_row_hash
            if not existing:
                existing = await db.staff_members.find_one({
                    "external_row_hash": row_hash,
                    "archived": {"$ne": True}
                })
            
            now = now_iso()
            
            staff_data = {
                "first_name": first_name,
                "last_name": last_name,
                "display_name": display_name,
                "email": email,
                "phone": phone,
                "phone_secondary": phone_secondary,
                "birthday": birthday,
                "personnel_number": personnel_number,
                "time_pin": time_pin,
                "address": address,
                "external_source": "excel_import",
                "external_row_hash": row_hash,
                "updated_at": now,
            }
            
            if existing:
                # Update existing
                await db.staff_members.update_one(
                    {"id": existing['id']},
                    {"$set": staff_data}
                )
                updated += 1
            else:
                # Create new
                staff_data["id"] = str(uuid.uuid4())
                staff_data["role"] = "service"  # Default
                staff_data["employment_type"] = "teilzeit"  # Default
                staff_data["weekly_hours"] = 0
                staff_data["hourly_rate"] = None
                staff_data["status"] = "aktiv"
                staff_data["work_areas"] = []
                staff_data["documents"] = []
                staff_data["created_at"] = now
                staff_data["archived"] = False
                
                await db.staff_members.insert_one(staff_data)
                imported += 1
                
        except Exception as e:
            logger.error(f"Error importing staff row {row_idx}: {e}")
            error_details.append({"row": row_idx, "reason": str(e)})
            errors += 1
    
    # Log import
    await log_import(
        user.get('email', 'unknown'),
        file.filename,
        'staff_members',
        imported,
        updated,
        errors
    )
    
    # Audit
    await create_audit_log(
        actor={"id": user['id'], "email": user['email']},
        entity="staff_members",
        entity_id="bulk_import",
        action="import",
        after={"filename": file.filename, "imported": imported, "updated": updated}
    )
    
    return {
        "success": errors == 0,
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "error_details": error_details[:10],  # Limit to first 10 errors
        "message": f"Import: {imported} neu, {updated} aktualisiert, {skipped} übersprungen, {errors} Fehler"
    }


@import_router.get("/api/admin/import/staff/preview")
async def preview_staff_import(user: dict = Depends(require_admin)):
    """
    GET /api/admin/import/staff/preview
    Show last imported staff members (first 20)
    """
    staff = await db.staff_members.find(
        {"external_source": "excel_import", "archived": {"$ne": True}},
        {"_id": 0, "id": 1, "first_name": 1, "last_name": 1, "email": 1, "phone": 1, 
         "personnel_number": 1, "status": 1, "created_at": 1}
    ).sort("created_at", -1).limit(20).to_list(20)
    
    return {
        "count": len(staff),
        "staff": staff
    }


@import_router.get("/api/admin/import/staff/status")
async def get_staff_import_status(user: dict = Depends(require_admin)):
    """
    GET /api/admin/import/staff/status
    Show last staff import status
    """
    # Get last staff import log
    last_import = await db.import_logs.find_one(
        {"collection": "staff_members"},
        sort=[("timestamp", -1)]
    )
    
    # Get counts
    total_staff = await db.staff_members.count_documents({"archived": {"$ne": True}})
    imported_staff = await db.staff_members.count_documents({
        "external_source": "excel_import",
        "archived": {"$ne": True}
    })
    
    return {
        "total_staff": total_staff,
        "imported_from_excel": imported_staff,
        "last_import": {
            "timestamp": last_import.get("timestamp") if last_import else None,
            "filename": last_import.get("filename") if last_import else None,
            "created": last_import.get("created") if last_import else 0,
            "updated": last_import.get("updated") if last_import else 0,
            "errors": last_import.get("errors") if last_import else 0,
        } if last_import else None
    }

        wb = openpyxl.load_workbook(tables_file)
        sheet_name = 'tables' if 'tables' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = [str(c.value).lower() if c.value else f"col{i}" for i, c in enumerate(ws[1])]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            try:
                data = dict(zip(headers, row))
                table_number = str(data.get('table_number', '')).strip()
                area = str(data.get('area', 'restaurant')).strip()
                subarea = data.get('subarea') or data.get('sub_area')
                if subarea:
                    subarea = str(subarea).strip()
                    if subarea == 'terrasse':
                        subarea = None
                
                existing = await db.tables.find_one({"table_number": table_number, "area": area})
                now = now_iso()
                
                if existing:
                    await db.tables.update_one(
                        {"id": existing['id']},
                        {"$set": {
                            "sub_area": subarea,
                            "seats_default": int(data.get('seats', 4)),
                            "seats_max": int(data.get('max_seats', data.get('seats', 4))),
                            "combinable": str(data.get('combinable', 'true')).lower() == 'true',
                            "notes": data.get('notes'),
                            "updated_at": now
                        }}
                    )
                    total_updated += 1
                else:
                    await db.tables.insert_one({
                        "id": str(uuid.uuid4()),
                        "table_number": table_number,
                        "area": area,
                        "sub_area": subarea,
                        "seats_default": int(data.get('seats', 4)),
                        "seats_max": int(data.get('max_seats', data.get('seats', 4))),
                        "combinable": str(data.get('combinable', 'true')).lower() == 'true',
                        "combinable_with": [],
                        "notes": data.get('notes'),
                        "active": True,
                        "fixed": False,
                        "created_at": now,
                        "updated_at": now,
                        "archived": False
                    })
                    total_created += 1
            except Exception as e:
                total_errors += 1
    
    # Import combinations
    if combos_file.exists():
        wb = openpyxl.load_workbook(combos_file)
        sheet_name = 'combinations' if 'combinations' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = [str(c.value).lower() if c.value else f"col{i}" for i, c in enumerate(ws[1])]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            try:
                data = dict(zip(headers, row))
                combo_id = str(data.get('combo_id', '')).strip()
                subarea = str(data.get('subarea', '')).strip()
                tables_str = str(data.get('tables', '')).strip()
                tables_list = [t.strip() for t in tables_str.split('+') if t.strip()]
                
                area = 'restaurant' if subarea in ['saal', 'wintergarten'] else subarea
                
                existing = await db.table_combinations.find_one({"combo_id": combo_id})
                now = now_iso()
                
                if existing:
                    await db.table_combinations.update_one(
                        {"id": existing['id']},
                        {"$set": {
                            "area": area,
                            "subarea": subarea,
                            "tables": tables_list,
                            "target_capacity": int(data.get('target_capacity', 0)) if data.get('target_capacity') else 0,
                            "notes": data.get('notes'),
                            "updated_at": now
                        }}
                    )
                    total_updated += 1
                else:
                    await db.table_combinations.insert_one({
                        "id": str(uuid.uuid4()),
                        "combo_id": combo_id,
                        "area": area,
                        "subarea": subarea,
                        "tables": tables_list,
                        "target_capacity": int(data.get('target_capacity', 0)) if data.get('target_capacity') else 0,
                        "notes": data.get('notes'),
                        "active": True,
                        "created_at": now,
                        "updated_at": now,
                        "archived": False
                    })
                    total_created += 1
            except Exception as e:
                total_errors += 1
    
    # Log
    await log_import(user.get('email', 'unknown'), "seed/from-repo", "tables+combinations", total_created, total_updated, total_errors)
    
    return {
        "success": total_errors == 0,
        "created": total_created,
        "updated": total_updated,
        "errors": total_errors,
        "message": f"Seed abgeschlossen: {total_created} neu, {total_updated} aktualisiert"
    }


@import_router.get("/api/admin/import/logs")
async def get_import_logs(limit: int = 20, user: dict = Depends(require_admin)):
    """Get recent import logs"""
    logs = await db.import_logs.find().sort("timestamp", -1).limit(limit).to_list(limit)
    for log in logs:
        log.pop('_id', None)
    return logs
