"""
GastroCore Backup & Export Module - Sprint: Admin Backup/Export
================================================================

FEATURES:
1. Backup Status anzeigen
2. Export XLSX (Staff + Tables)
3. Export JSON (Events/Actions)
4. Write Backup to Disk

SECURITY:
- Admin only
- Sensitive fields masked
- No secrets in exports

ADDITIV - Keine Breaking Changes
"""

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
from pathlib import Path
import uuid
import io
import json
import os
import logging

# Core imports
from core.database import db
from core.auth import require_admin
from core.audit import create_audit_log, SYSTEM_ACTOR

logger = logging.getLogger(__name__)

# ============== CONFIG ==============
BACKUP_FOLDER = Path("/app/backups")
BACKUP_FOLDER.mkdir(parents=True, exist_ok=True)

# Router
backup_router = APIRouter(prefix="/api/admin/backup", tags=["Backup & Export"])


# ============== PYDANTIC MODELS ==============
class BackupFile(BaseModel):
    name: str
    size: int
    modified: str


class BackupStatusResponse(BaseModel):
    enabled: bool
    last_backup_at: Optional[str]
    backup_folder: str
    files: List[BackupFile]


class WriteBackupResponse(BaseModel):
    ok: bool
    written: List[str]
    folder: str


# ============== HELPER FUNCTIONS ==============
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def mask_sensitive(value: str, visible_chars: int = 4) -> str:
    """Mask sensitive data, showing only last N characters"""
    if not value:
        return ""
    if len(value) <= visible_chars:
        return "*" * len(value)
    return "*" * (len(value) - visible_chars) + value[-visible_chars:]


def get_backup_files() -> List[Dict]:
    """Get list of files in backup folder"""
    files = []
    if BACKUP_FOLDER.exists():
        for f in BACKUP_FOLDER.iterdir():
            if f.is_file() and not f.name.startswith('.'):
                stat = f.stat()
                files.append({
                    "name": f.name,
                    "size": stat.st_size,
                    "modified": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
                })
    return sorted(files, key=lambda x: x['modified'], reverse=True)


def get_last_backup_time() -> Optional[str]:
    """Get timestamp of most recent backup"""
    files = get_backup_files()
    if files:
        return files[0]['modified']
    return None


# ============== ENDPOINTS ==============

@backup_router.get("/status", response_model=BackupStatusResponse)
async def get_backup_status(user: dict = Depends(require_admin)):
    """
    GET /api/admin/backup/status
    Returns backup system status and list of existing backup files
    """
    return {
        "enabled": True,
        "last_backup_at": get_last_backup_time(),
        "backup_folder": str(BACKUP_FOLDER),
        "files": get_backup_files()
    }


@backup_router.get("/export-xlsx")
async def export_xlsx(user: dict = Depends(require_admin)):
    """
    GET /api/admin/backup/export-xlsx
    Download XLSX with staff_members and tables data
    Sensitive fields are masked
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    wb = Workbook()
    
    # ===== Sheet 1: Staff Members =====
    ws_staff = wb.active
    ws_staff.title = "staff_members"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002f02", end_color="002f02", fill_type="solid")
    
    staff_headers = [
        "ID", "Vorname", "Nachname", "Rufname", "E-Mail", "Telefon",
        "Pers-Nr.", "Zeit-PIN", "Geburtstag", "Rolle", "Beschäftigung",
        "Status", "Arbeitsbereiche", "Adresse", "Stadt", "PLZ",
        "Steuer-ID (maskiert)", "SV-Nr. (maskiert)", "IBAN (maskiert)",
        "Erstellt am"
    ]
    
    for col, header in enumerate(staff_headers, 1):
        cell = ws_staff.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    staff_members = await db.staff_members.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(1000)
    
    for row_idx, staff in enumerate(staff_members, 2):
        address = staff.get('address', {}) or {}
        ws_staff.cell(row=row_idx, column=1, value=staff.get('id', ''))
        ws_staff.cell(row=row_idx, column=2, value=staff.get('first_name', ''))
        ws_staff.cell(row=row_idx, column=3, value=staff.get('last_name', ''))
        ws_staff.cell(row=row_idx, column=4, value=staff.get('display_name', ''))
        ws_staff.cell(row=row_idx, column=5, value=staff.get('email', ''))
        ws_staff.cell(row=row_idx, column=6, value=staff.get('phone', ''))
        ws_staff.cell(row=row_idx, column=7, value=staff.get('personnel_number', ''))
        ws_staff.cell(row=row_idx, column=8, value=staff.get('time_pin', ''))
        ws_staff.cell(row=row_idx, column=9, value=staff.get('birthday', ''))
        ws_staff.cell(row=row_idx, column=10, value=staff.get('role', ''))
        ws_staff.cell(row=row_idx, column=11, value=staff.get('employment_type', ''))
        ws_staff.cell(row=row_idx, column=12, value=staff.get('status', ''))
        ws_staff.cell(row=row_idx, column=13, value=', '.join(staff.get('work_areas', [])))
        ws_staff.cell(row=row_idx, column=14, value=address.get('street', ''))
        ws_staff.cell(row=row_idx, column=15, value=address.get('city', ''))
        ws_staff.cell(row=row_idx, column=16, value=address.get('zip', ''))
        # Masked sensitive fields
        ws_staff.cell(row=row_idx, column=17, value=mask_sensitive(staff.get('tax_id', ''), 4))
        ws_staff.cell(row=row_idx, column=18, value=mask_sensitive(staff.get('social_security_number', ''), 4))
        ws_staff.cell(row=row_idx, column=19, value=mask_sensitive(staff.get('bank_iban', ''), 4))
        ws_staff.cell(row=row_idx, column=20, value=staff.get('created_at', ''))
    
    # Adjust column widths
    for col in ws_staff.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_staff.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # ===== Sheet 2: Tables =====
    ws_tables = wb.create_sheet("tables")
    
    table_headers = [
        "ID", "Tischnummer", "Bereich", "Subbereich", "Plätze (max)",
        "Plätze (default)", "Kombinierbar", "Kombinierbar mit",
        "Fest", "Aktiv", "Position X", "Position Y", "Notizen", "Erstellt am"
    ]
    
    for col, header in enumerate(table_headers, 1):
        cell = ws_tables.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    tables = await db.tables.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(500)
    
    for row_idx, table in enumerate(tables, 2):
        ws_tables.cell(row=row_idx, column=1, value=table.get('id', ''))
        ws_tables.cell(row=row_idx, column=2, value=table.get('table_number', ''))
        ws_tables.cell(row=row_idx, column=3, value=table.get('area', ''))
        ws_tables.cell(row=row_idx, column=4, value=table.get('sub_area', ''))
        ws_tables.cell(row=row_idx, column=5, value=table.get('seats_max', ''))
        ws_tables.cell(row=row_idx, column=6, value=table.get('seats_default', ''))
        ws_tables.cell(row=row_idx, column=7, value="Ja" if table.get('combinable') else "Nein")
        ws_tables.cell(row=row_idx, column=8, value=', '.join(table.get('combinable_with', [])))
        ws_tables.cell(row=row_idx, column=9, value="Ja" if table.get('fixed') else "Nein")
        ws_tables.cell(row=row_idx, column=10, value="Ja" if table.get('active') else "Nein")
        ws_tables.cell(row=row_idx, column=11, value=table.get('position_x', ''))
        ws_tables.cell(row=row_idx, column=12, value=table.get('position_y', ''))
        ws_tables.cell(row=row_idx, column=13, value=table.get('notes', ''))
        ws_tables.cell(row=row_idx, column=14, value=table.get('created_at', ''))
    
    # Adjust column widths
    for col in ws_tables.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_tables.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    # ===== Sheet 3: Meta =====
    ws_meta = wb.create_sheet("meta")
    ws_meta.cell(row=1, column=1, value="Export-Zeitpunkt")
    ws_meta.cell(row=1, column=2, value=now_iso())
    ws_meta.cell(row=2, column=1, value="Exportiert von")
    ws_meta.cell(row=2, column=2, value=user.get('email', 'unknown'))
    ws_meta.cell(row=3, column=1, value="Staff Members")
    ws_meta.cell(row=3, column=2, value=len(staff_members))
    ws_meta.cell(row=4, column=1, value="Tables")
    ws_meta.cell(row=4, column=2, value=len(tables))
    ws_meta.cell(row=5, column=1, value="Hinweis")
    ws_meta.cell(row=5, column=2, value="Sensible Daten (Steuer-ID, SV-Nr., IBAN) sind maskiert.")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"Carlsburg_Backup_{timestamp}.xlsx"
    
    # Audit log
    await create_audit_log(
        actor={"id": user['id'], "email": user['email']},
        action="backup_export_xlsx",
        entity_type="backup",
        entity_id=filename,
        new_value={"staff_count": len(staff_members), "tables_count": len(tables)}
    )
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@backup_router.get("/export-events-actions-json")
async def export_events_actions_json(user: dict = Depends(require_admin)):
    """
    GET /api/admin/backup/export-events-actions-json
    Download JSON with events and actions data
    """
    events = await db.events.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(1000)
    actions = await db.actions.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(1000)
    
    # Also get menu_actions if separate collection exists
    menu_actions = []
    try:
        menu_actions = await db.menu_actions.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(500)
    except:
        pass
    
    data = {
        "export_timestamp": now_iso(),
        "exported_by": user.get('email', 'unknown'),
        "events": events,
        "actions": actions,
        "menu_actions": menu_actions,
        "counts": {
            "events": len(events),
            "actions": len(actions),
            "menu_actions": len(menu_actions)
        }
    }
    
    # Convert to JSON bytes
    json_bytes = json.dumps(data, indent=2, ensure_ascii=False, default=str).encode('utf-8')
    output = io.BytesIO(json_bytes)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"Carlsburg_EventsActions_{timestamp}.json"
    
    # Audit log
    await create_audit_log(
        actor={"id": user['id'], "email": user['email']},
        action="backup_export_events_json",
        entity_type="backup",
        entity_id=filename,
        new_value={"events_count": len(events), "actions_count": len(actions)}
    )
    
    return StreamingResponse(
        output,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@backup_router.post("/write", response_model=WriteBackupResponse)
async def write_backup_to_disk(user: dict = Depends(require_admin)):
    """
    POST /api/admin/backup/write
    Write backup files to /app/backups/ folder
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    
    written_files = []
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    
    # Ensure backup folder exists
    BACKUP_FOLDER.mkdir(parents=True, exist_ok=True)
    
    # 1. Write XLSX Backup
    wb = Workbook()
    ws_staff = wb.active
    ws_staff.title = "staff_members"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002f02", end_color="002f02", fill_type="solid")
    
    # Staff sheet
    staff_headers = ["ID", "Vorname", "Nachname", "Rufname", "E-Mail", "Telefon",
                     "Pers-Nr.", "Zeit-PIN", "Geburtstag", "Rolle", "Status"]
    for col, header in enumerate(staff_headers, 1):
        cell = ws_staff.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    staff_members = await db.staff_members.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(1000)
    for row_idx, staff in enumerate(staff_members, 2):
        ws_staff.cell(row=row_idx, column=1, value=staff.get('id', ''))
        ws_staff.cell(row=row_idx, column=2, value=staff.get('first_name', ''))
        ws_staff.cell(row=row_idx, column=3, value=staff.get('last_name', ''))
        ws_staff.cell(row=row_idx, column=4, value=staff.get('display_name', ''))
        ws_staff.cell(row=row_idx, column=5, value=staff.get('email', ''))
        ws_staff.cell(row=row_idx, column=6, value=staff.get('phone', ''))
        ws_staff.cell(row=row_idx, column=7, value=staff.get('personnel_number', ''))
        ws_staff.cell(row=row_idx, column=8, value=staff.get('time_pin', ''))
        ws_staff.cell(row=row_idx, column=9, value=staff.get('birthday', ''))
        ws_staff.cell(row=row_idx, column=10, value=staff.get('role', ''))
        ws_staff.cell(row=row_idx, column=11, value=staff.get('status', ''))
    
    # Tables sheet
    ws_tables = wb.create_sheet("tables")
    table_headers = ["ID", "Tischnummer", "Bereich", "Subbereich", "Plätze", "Kombinierbar", "Aktiv"]
    for col, header in enumerate(table_headers, 1):
        cell = ws_tables.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    tables = await db.tables.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(500)
    for row_idx, table in enumerate(tables, 2):
        ws_tables.cell(row=row_idx, column=1, value=table.get('id', ''))
        ws_tables.cell(row=row_idx, column=2, value=table.get('table_number', ''))
        ws_tables.cell(row=row_idx, column=3, value=table.get('area', ''))
        ws_tables.cell(row=row_idx, column=4, value=table.get('sub_area', ''))
        ws_tables.cell(row=row_idx, column=5, value=table.get('seats_max', ''))
        ws_tables.cell(row=row_idx, column=6, value="Ja" if table.get('combinable') else "Nein")
        ws_tables.cell(row=row_idx, column=7, value="Ja" if table.get('active') else "Nein")
    
    xlsx_filename = f"Carlsburg_Backup_{timestamp}.xlsx"
    xlsx_path = BACKUP_FOLDER / xlsx_filename
    wb.save(xlsx_path)
    written_files.append(xlsx_filename)
    
    # 2. Write Events/Actions JSON
    events = await db.events.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(1000)
    actions = await db.actions.find({"archived": {"$ne": True}}, {"_id": 0}).to_list(1000)
    
    json_data = {
        "export_timestamp": now_iso(),
        "exported_by": user.get('email', 'unknown'),
        "events": events,
        "actions": actions,
        "counts": {"events": len(events), "actions": len(actions)}
    }
    
    json_filename = f"Carlsburg_EventsActions_{timestamp}.json"
    json_path = BACKUP_FOLDER / json_filename
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
    written_files.append(json_filename)
    
    # Audit log
    await create_audit_log(
        actor={"id": user['id'], "email": user['email']},
        action="backup_write_disk",
        entity_type="backup",
        entity_id=timestamp,
        new_value={"files": written_files, "folder": str(BACKUP_FOLDER)}
    )
    
    logger.info(f"Backup written to {BACKUP_FOLDER}: {written_files}")
    
    return {
        "ok": True,
        "written": written_files,
        "folder": str(BACKUP_FOLDER)
    }
